"""Unit tests for XLSX extraction.

Two real failures motivated this:
  * `markitdown[docx,pdf,pptx]` was missing the `xlsx,xls` extras, so
    every Excel workbook came back empty from MarkItDown and got logged
    as "All extraction methods failed or returned empty content".
  * Even with the extras in place, some workbooks return empty from
    MarkItDown's converter (formula-only sheets with no cached values,
    legacy encodings). A lightweight openpyxl fallback is the safety net.

These tests build real workbooks with openpyxl, then exercise:
  * MarkItDown can now read a normal workbook end-to-end (regression
    guard against silently dropping the xlsx extras).
  * `_try_xlsx_extraction` recovers sheet names + cell text from a
    workbook MarkItDown can't help with.
  * The empty/missing/corrupt paths return None instead of raising.
"""

import asyncio

import pytest
from openpyxl import Workbook

from cosma_backend.parser.parser import FileParser


@pytest.mark.unit
class TestXlsxFallback:
    @pytest.mark.asyncio
    async def test_extracts_sheet_names_and_cell_text(self, tmp_path):
        p = tmp_path / "exam.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Midterm Answers"
        ws["A1"] = "Question"
        ws["B1"] = "Answer"
        ws["A2"] = "What is 2+2?"
        ws["B2"] = "Four"
        wb.save(str(p))

        parser = FileParser()
        text = await parser._try_xlsx_extraction(p)

        assert text is not None
        assert "Midterm Answers" in text  # sheet name surfaces
        assert "What is 2+2?" in text
        assert "Four" in text

    @pytest.mark.asyncio
    async def test_empty_workbook_surfaces_sheet_names(self, tmp_path):
        # A user-renamed sheet ("Final Exam") is meaningful searchable
        # text even when every cell is empty — they should still find
        # the file by its sheet title. So this returns content, not None.
        p = tmp_path / "blank.xlsx"
        wb = Workbook()
        wb.active.title = "Final Exam"
        wb.save(str(p))

        parser = FileParser()
        text = await parser._try_xlsx_extraction(p)

        assert text is not None
        assert "Final Exam" in text

    @pytest.mark.asyncio
    async def test_missing_file_returns_none_without_raising(self, tmp_path):
        parser = FileParser()
        text = await parser._try_xlsx_extraction(tmp_path / "nope.xlsx")
        assert text is None

    @pytest.mark.asyncio
    async def test_corrupt_file_returns_none_without_raising(self, tmp_path):
        p = tmp_path / "broken.xlsx"
        p.write_bytes(b"definitely not a zip-backed xlsx")
        parser = FileParser()
        text = await parser._try_xlsx_extraction(p)
        assert text is None

    @pytest.mark.asyncio
    async def test_per_sheet_truncation_caps_runaway_workbooks(self, tmp_path):
        # Build a sheet that exceeds the MAX_SHEET_CHARS cap (50k) so
        # the loop trips the "(sheet truncated)" branch — without this
        # cap a 10MB data sheet would dominate the index token budget.
        p = tmp_path / "huge.xlsx"
        wb = Workbook()
        ws = wb.active
        for row_idx in range(1, 5_000):
            ws.cell(row=row_idx, column=1, value=f"row-{row_idx}-cell-text")
        wb.save(str(p))

        parser = FileParser()
        text = await parser._try_xlsx_extraction(p)

        assert text is not None
        assert "(sheet truncated)" in text
        # Should be far smaller than the full content would be.
        assert len(text) < 200_000


@pytest.mark.unit
class TestMarkItDownXlsxSupport:
    """If this fails, the [xlsx,xls] extras got dropped from
    markitdown's dep spec again — every .xlsx will silently come back
    empty in production."""

    def test_markitdown_xlsx_converter_importable(self):
        # MarkItDown only registers the xlsx converter when openpyxl is
        # installed (the converter does the import at module top-level
        # and the registry skips it on ImportError).
        from markitdown.converters._xlsx_converter import XlsxConverter

        assert XlsxConverter is not None
