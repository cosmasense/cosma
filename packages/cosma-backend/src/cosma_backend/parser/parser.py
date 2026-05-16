from __future__ import annotations

import asyncio
import hashlib
import mimetypes
import os
import zipfile
from pathlib import Path
from typing import Optional, Dict, Any, TYPE_CHECKING

from markitdown import MarkItDown

from cosma_backend.dev_logging import log_failed_file
from cosma_backend.logging import get_logger
from cosma_backend.models import File, ProcessingStatus
from cosma_backend.parser.spotlight import spotlight_to_text
from cosma_backend.parser.media import (
    extract_audio_tags,
    extract_audio_transcript,
    extract_video_content,
    extract_video_transcript,
    extract_image_info,
    is_supported_media_file
)

if TYPE_CHECKING:
    from cosma_backend.settings import ParserConfig

# Configure structured logger
logger = get_logger(__name__)


class FileParser:
    """
    High-performance file parser that converts various file formats to markdown.

    Uses Microsoft's MarkItDown library to support 20+ file formats including:
    - Documents: PDF, DOCX, PPTX, XLSX
    - Images: PNG, JPG, JPEG, TIFF, BMP, HEIC (with OCR)
    - Audio: MP3, WAV, AAC, M4A, FLAC, OGG/Opus, WMA, AIFF, ... (transcribed
      via whisper after an ffmpeg normalize-to-PCM pass)
    - Video: MP4, MOV, MKV, AVI, WEBM, ... (frames + transcript)
    - Web: HTML pages
    - Text: CSV, JSON, XML, TXT
    - Archives: ZIP (processes contents)
    - And more formats...
    """

    # Supported file extensions (MarkItDown supports many more)
    SUPPORTED_EXTENSIONS = {
        # Documents
        ".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls",
        # Images
        ".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".heic", ".gif", ".webp",
        # Audio — anything ffmpeg can decode; we normalize to 16 kHz mono
        # PCM before whisper, so the container/codec just has to be
        # ffmpeg-readable. Keep this in sync with
        # media.get_supported_media_extensions()["audio"].
        ".mp3", ".wav", ".aac", ".m4a", ".m4b", ".flac", ".ogg",
        ".opus", ".wma", ".aiff", ".aif", ".alac", ".amr", ".ac3",
        # Video
        ".mp4", ".avi", ".mov", ".mkv", ".wmv", ".flv", ".webm",
        # Web & Text
        ".html", ".htm", ".txt", ".csv", ".json", ".xml", ".md",
        # Archives
        ".zip", ".epub"
    }

    def __init__(self, config: ParserConfig | None = None) -> None:
        """Initialize the parser with MarkItDown."""
        from cosma_backend.settings import ParserConfig as _ParserConfig
        self.config = config or _ParserConfig()
        try:
            self.markitdown = MarkItDown()

            # Extraction strategy configuration
            self.extraction_strategy = self.config.extraction_strategy
            self.spotlight_enabled = self.config.spotlight_enabled
            
            # Statistics tracking
            self.extraction_stats = {
                "spotlight_success": 0,
                "spotlight_failed": 0,
                "markitdown_success": 0,
                "markitdown_failed": 0,
                "media_success": 0,
                "media_failed": 0
            }
            
            logger.info("FileParser initialized successfully",
                          extraction_strategy=self.extraction_strategy,
                          spotlight_enabled=self.spotlight_enabled)
        except Exception as e:
            logger.exception("Failed to initialize MarkItDown", error=str(e))
            raise

    def is_supported(self, file: File) -> bool:
        """
        Check if a file format is supported by the parser.

        Args:
            file_path: Path to the file to check

        Returns:
            True if the file format is supported, False otherwise
        """
        extension = file.path.suffix.lower()
        return extension in self.SUPPORTED_EXTENSIONS

    def get_supported_extensions(self) -> set[str]:
        """Get the set of supported file extensions."""
        return self.SUPPORTED_EXTENSIONS.copy()

    def _calculate_content_hash(self, content: str) -> str:
        """Calculate SHA-256 hash of content for deduplication."""
        return hashlib.sha256(content.encode("utf-8")).hexdigest()

    def _detect_mime_type(self, file_path: Path) -> str | None:
        """Detect MIME type of a file."""
        try:
            mime_type, _ = mimetypes.guess_type(str(file_path))
            return mime_type
        except Exception:
            return None

    async def parse_file(self, file: File) -> File:
        """
        Parse a single file using hybrid extraction strategy (Phase 2).
        
        Extraction order:
        1. Spotlight extraction (macOS) - if enabled and available
        2. Media processing (audio/video/image) - if media file
        3. MarkItDown fallback - for all other supported formats

        Args:
            file_path: Path to the file to parse
            title: Optional custom title for the file

        Returns:
            File object with parsed content

        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is not supported
            Exception: For other parsing errors
        """
        logger.info("Starting hybrid file parsing", 
                       file_path=str(file.file_path), 
                       title=file.title,
                       strategy=self.extraction_strategy)
        
        path = Path(file.file_path)

        # Validate file exists
        if not path.exists():
            error_msg = f"File not found: {file.file_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        # Check if format is supported
        if not self.is_supported(file):
            error_msg = f"Unsupported file format: {path.suffix}"
            logger.error(error_msg, extension=path.suffix)
            raise ValueError(error_msg)
            
        max_file_size = self.config.max_file_size_mb * 1024 * 1024
        if file.file_size > max_file_size:
            # Don't hard-fail: route through the metadata-only path so the
            # file is still findable by name (a 60-GB movie remux the user
            # searches for by title should turn up, even though we won't
            # transcode or OCR it). It still ends in status=FAILED and
            # shows in the Failed tab — the user asked us to keep surfacing
            # these — but now with a useful "indexed by filename only"
            # reason instead of vanishing.
            error_msg = (
                f"File exceeds the {self.config.max_file_size_mb}MB parse "
                f"limit; indexed by filename only"
            )
            logger.warning(error_msg, file_path=str(path),
                           size=file.file_size, max=max_file_size)
            file.content_type = self._detect_mime_type(path)
            file.content = _generic_metadata_placeholder(
                path, f"exceeds the {self.config.max_file_size_mb}MB parse limit"
            )
            file.metadata_only = True
            if file.partial_kind is None:
                file.partial_kind = "parser_failed"
            file.metadata_only_reason = error_msg
            log_failed_file(
                file_path=str(path),
                extension=path.suffix.lower(),
                error=error_msg,
                phase="parse",
            )
            return file


        try:
            # Detect and set MIME type
            file.content_type = self._detect_mime_type(path)

            # Phase 2: Hybrid extraction strategy
            extracted_content = None
            extraction_method = None
            
            # Strategy 1: Try Spotlight extraction first (if enabled and strategy allows)
            if (self.spotlight_enabled and 
                self.extraction_strategy in ["spotlight_first"]):
                
                extracted_content = await self._try_spotlight_extraction(path)
                if extracted_content:
                    extraction_method = "spotlight"
                    self.extraction_stats["spotlight_success"] += 1
                    logger.info("EXTRACTION: Spotlight successful", 
                                   filename=path.name,
                                   method="spotlight", 
                                   content_length=len(extracted_content))
                else:
                    self.extraction_stats["spotlight_failed"] += 1
                    logger.info("EXTRACTION: Spotlight failed, falling back",
                                   filename=path.name,
                                   method="spotlight_failed")

            # Strategy 2: Try media processing for audio/video/image files
            if not extracted_content:
                is_media, media_type = await is_supported_media_file(path)
                if is_media:
                    extracted_content = await self._try_media_extraction(path, media_type, file)
                    if extracted_content:
                        extraction_method = f"media_{media_type}"
                        self.extraction_stats["media_success"] += 1
                        logger.info("EXTRACTION: Media successful",
                                       filename=path.name,
                                       media_type=media_type,
                                       method=f"media_{media_type}", 
                                       content_length=len(extracted_content))
                    else:
                        self.extraction_stats["media_failed"] += 1
                        logger.info("EXTRACTION: Media failed",
                                       filename=path.name,
                                       media_type=media_type,
                                       method=f"media_{media_type}_failed")

            # Strategy 3: Fallback to MarkItDown
            if not extracted_content and self.extraction_strategy != "spotlight_only":
                extracted_content = await self._try_markitdown_extraction(path)
                if extracted_content:
                    extraction_method = "markitdown"
                    self.extraction_stats["markitdown_success"] += 1
                    logger.info("EXTRACTION: MarkItDown successful",
                                   filename=path.name,
                                   method="markitdown",
                                   content_length=len(extracted_content))
                else:
                    self.extraction_stats["markitdown_failed"] += 1
                    logger.info("EXTRACTION: MarkItDown failed",
                                   filename=path.name,
                                   method="markitdown_failed")

            # Strategy 4: Vision OCR — last-resort for scanned PDFs and
            # raster images whose text content every prior method gave up on.
            # Why this is gated on PDF ext: OCR is expensive (~1s/page on
            # M-series) and most other formats that reach this point are
            # binaries with genuinely no text, not scanned documents.
            # The Carol-frontend failure report had at least one scanned
            # PDF (`wechat-channels-en.pdf`) that sat in this exact hole.
            if not extracted_content and path.suffix.lower() == ".pdf":
                from cosma_backend.parser import ocr_vision
                if ocr_vision.is_available():
                    logger.info("Attempting Vision OCR fallback", filename=path.name)
                    ocr_text = await ocr_vision.ocr_pdf(path)
                    if ocr_text:
                        extracted_content = ocr_text
                        extraction_method = "vision_ocr"
                        logger.info("EXTRACTION: Vision OCR successful",
                                    filename=path.name,
                                    content_length=len(ocr_text))

            # Strategy 5: PDF text via pypdf — MarkItDown sometimes returns
            # empty for PDFs whose text layer is unusual (encrypted, weird
            # encodings) but pypdf can still scrape something. Cheap and
            # already pulled in by MarkItDown[pdf], so no new dep.
            if not extracted_content and path.suffix.lower() == ".pdf":
                pypdf_text = await self._try_pypdf_extraction(path)
                if pypdf_text:
                    extracted_content = pypdf_text
                    extraction_method = "pypdf"
                    logger.info("EXTRACTION: pypdf fallback successful",
                                filename=path.name,
                                content_length=len(pypdf_text))

            # Strategy 5b: XLSX fallback via openpyxl. Some workbooks return
            # empty from MarkItDown's converter (legacy encodings, malformed
            # shared-strings table, password-locked-but-readable, sheets that
            # are 100% formulas with no cached values). Lists sheet names +
            # any non-empty cell text — at minimum the file becomes findable
            # by sheet titles. openpyxl ships with markitdown[xlsx].
            if not extracted_content and path.suffix.lower() in (".xlsx", ".xlsm"):
                xlsx_text = await self._try_xlsx_extraction(path)
                if xlsx_text:
                    extracted_content = xlsx_text
                    extraction_method = "openpyxl"
                    logger.info("EXTRACTION: openpyxl fallback successful",
                                filename=path.name,
                                content_length=len(xlsx_text))

            # Strategy 6: ZIP listing — when MarkItDown can't text-extract
            # an archive (binary app/font bundles), enumerate entry names
            # and pull readable text from any inner README/.txt/.md/.json.
            # The entry list itself is meaningful searchable text — a user
            # looking for "VSCode" should find VSCode-darwin-universal.zip
            # by its inner Code Helper paths.
            if not extracted_content and path.suffix.lower() == ".zip":
                zip_text = await self._try_zip_listing_extraction(path)
                if zip_text:
                    extracted_content = zip_text
                    extraction_method = "zip_listing"
                    logger.info("EXTRACTION: ZIP listing successful",
                                filename=path.name,
                                content_length=len(zip_text))

            # Process results
            if extracted_content and len(extracted_content.strip()) > 10:
                file.content = extracted_content

                # Calculate content hash for deduplication
                content_hash = self._calculate_content_hash(extracted_content)
                file.content_hash = content_hash

                # Mark as successfully processed
                file.status = ProcessingStatus.PARSED

                logger.info("File parsed successfully",
                               filename=file.filename,
                               content_length=len(extracted_content),
                               extraction_method=extraction_method,
                               content_hash=content_hash[:12])

                return file
            else:
                # Every parser came back empty. Don't raise — instead route
                # through the existing metadata-only path so the file:
                #   * still ends in status=FAILED (visible in Failed tab,
                #     so the user — and we — see the failure),
                #   * is still embedded by filename + path + size, so it
                #     remains searchable by name,
                #   * carries a structured error message for telemetry.
                # The pipeline (pipeline.py:443) handles the FAILED+embed
                # routing once metadata_only=True is set.
                #
                # The reason string is honest about which case this is:
                # blank documents ("Blank 3.pdf") aren't really failures
                # — the extraction succeeded, the document just had no
                # content. The user sees these in the Failed tab today
                # alongside genuine crashes, which is misleading.
                error_msg = (
                    "No extractable text or images "
                    "(document appears empty or unreadable); "
                    "indexed by filename only"
                )
                logger.warning(error_msg, file_path=str(path))
                file.content = _generic_metadata_placeholder(path, error_msg)
                file.metadata_only = True
                if file.partial_kind is None:
                    file.partial_kind = "parser_failed"
                file.metadata_only_reason = error_msg
                # Local failure telemetry — no-op outside COSMA_DEV. Once
                # we have a remote endpoint, this is the single chokepoint
                # to add the network call on top of.
                log_failed_file(
                    file_path=str(path),
                    extension=path.suffix.lower(),
                    error=error_msg,
                    phase="parse",
                )
                return file

        except Exception as e:
            error_msg = f"Failed to parse file: {e!s}"
            logger.exception(error_msg, file_path=str(path), error=str(e))

            # Ensure file has a consistent failed state before re-raising
            file.status = ProcessingStatus.FAILED
            file.processing_error = error_msg
            raise

    async def _try_spotlight_extraction(self, path: Path) -> str | None:
        """
        Try extracting text using macOS Spotlight.
        
        Args:
            path: Path to file
            
        Returns:
            Extracted text or None if failed
        """
        try:
            content = await spotlight_to_text(path, self.config)
            if content and len(content.strip()) > 50:  # Minimum content threshold
                logger.debug("Spotlight extraction successful", 
                                path=str(path), 
                                length=len(content))
                return content
            else:
                logger.debug("Spotlight extraction returned insufficient content", 
                                path=str(path))
                return None
        except Exception as e:
            logger.debug("Spotlight extraction failed", 
                            path=str(path), 
                            error=str(e))
            return None

    async def _try_media_extraction(self, path: Path, media_type: str, file: File | None = None) -> str | None:
        """
        Try extracting content from media files.

        Args:
            path: Path to media file
            media_type: Type of media ('audio', 'video', 'image')
            file: Optional File object to attach extra data (e.g., video frames)

        Returns:
            Extracted content or None if failed
        """
        try:
            if media_type == "audio":
                content = await extract_audio_transcript(path, self.config)
                # No decodable speech — try embedded tags before giving up.
                # Most downloaded music has artist/title/album baked into
                # ID3/MP4 atoms, which is the actual signal a user wants
                # to search by. When tags are present we treat the file as
                # PARSED (not metadata-only): the search hit on "Charlie
                # Clouser Dead Silence" is real, not a fallback.
                if not content or len(content.strip()) <= 10:
                    tag_text = extract_audio_tags(path)
                    if tag_text:
                        content = tag_text
                        logger.info("Audio: using embedded tags as content",
                                    filename=path.name,
                                    tag_length=len(tag_text))
                    else:
                        # Truly nothing — silent/untagged track. Fall through
                        # to the metadata-only placeholder so the pipeline
                        # routes it to FAILED + embed (still findable by
                        # filename, no LLM summary, visible in Failed tab).
                        content = _audio_metadata_placeholder(path)
                        if file is not None:
                            file.metadata_only = True
                            # Don't overwrite a user_elected classification —
                            # parser-failure only claims partial_kind when
                            # nothing else has.
                            if file.partial_kind is None:
                                file.partial_kind = "parser_failed"
                            file.metadata_only_reason = (
                                "Audio: no usable speech transcript and no "
                                "embedded tags (silent/untagged track, codec "
                                "issue, or whisper unavailable)"
                            )
            elif media_type == "video":
                # Extract both transcript and frames for vision analysis
                video_content = await extract_video_content(path)
                content = video_content.transcript
                # Attach frames to file for summarizer vision analysis
                if file is not None and video_content.frames:
                    file.extra_images = video_content.frames
                    logger.debug("Attached video frames for vision analysis",
                                 path=str(path),
                                 num_frames=len(video_content.frames))
                    # Silent video with frames — still PARSED success: the
                    # summarizer's vision pass will run against the frames.
                    if not content or len(content.strip()) <= 10:
                        content = (
                            f"Video file: {path.name} "
                            f"({len(video_content.frames)} frames extracted "
                            f"for visual description)."
                        )
                # Last resort: ffmpeg couldn't decode frames *and* there's
                # no transcript. We still surface metadata so the file is
                # findable by name, but flag metadata_only so the pipeline
                # routes it to FAILED + embed-only (no LLM summary on the
                # synthetic placeholder text).
                if not content or len(content.strip()) <= 10:
                    content = _video_metadata_placeholder(path)
                    if file is not None:
                        file.metadata_only = True
                        if file.partial_kind is None:
                            file.partial_kind = "parser_failed"
                        file.metadata_only_reason = (
                            "Video: no transcript or frames extracted "
                            "(codec unreadable even after H.264 transcode "
                            "fallback, or ffmpeg unavailable)"
                        )
            elif media_type == "image":
                # For images, we'll let the summarization process handle vision analysis
                # Just extract basic image info here
                image_info = await extract_image_info(path)
                if image_info:
                    dimensions_str = ""
                    if image_info.get("dimensions"):
                        dims = image_info["dimensions"]
                        dimensions_str = f" ({dims['width']}x{dims['height']}, {dims.get('mode', 'unknown')} mode)"
                    
                    content = f"Image file: {path.name}{dimensions_str}, Size: {image_info['file_size']} bytes, Type: {image_info['file_type']}"
                else:
                    content = f"Image file: {path.name}"
            else:
                logger.warning("Unknown media type", media_type=media_type)
                return None
                
            if content and len(content.strip()) > 10:
                logger.debug("Media extraction successful", 
                                path=str(path), 
                                media_type=media_type,
                                length=len(content))
                return content
            else:
                logger.debug("Media extraction returned insufficient content", 
                                path=str(path), 
                                media_type=media_type)
                return None
                
        except Exception as e:
            logger.debug("Media extraction failed", 
                            path=str(path), 
                            media_type=media_type,
                            error=str(e))
            return None

    async def _try_markitdown_extraction(self, path: Path) -> str | None:
        """
        Try extracting content using MarkItDown.
        
        Args:
            path: Path to file
            
        Returns:
            Extracted markdown content or None if failed
        """
        try:
            logger.debug("Trying MarkItDown extraction", path=str(path))

            # Timeout prevents hangs on malformed/malicious files
            from cosma_backend.pipeline_executor import run_in_pipeline
            result = await asyncio.wait_for(
                run_in_pipeline(self.markitdown.convert, str(path)),
                timeout=120,  # 2 minute max per file
            )

            if result and result.text_content:
                content = result.text_content.strip()
                if len(content) > 0:
                    logger.debug("MarkItDown extraction successful", 
                                    path=str(path), 
                                    length=len(content))
                    return content
                    
            logger.debug("MarkItDown extraction returned empty content", path=str(path))
            return None
            
        except asyncio.TimeoutError:
            logger.warning("MarkItDown extraction timed out", path=str(path))
            return None
        except Exception as e:
            logger.debug("MarkItDown extraction failed",
                            path=str(path),
                            error=str(e))
            return None

    async def _try_pypdf_extraction(self, path: Path) -> str | None:
        """Last-resort PDF text scrape via pypdf. MarkItDown's pdfminer
        path silently returns empty for some PDFs (encrypted with empty
        password, unusual encodings); pypdf occasionally salvages those.
        Returns None on any failure — caller decides what to do next.
        pypdf comes in via MarkItDown[pdf], no new dep.
        """
        try:
            from pypdf import PdfReader
            from cosma_backend.pipeline_executor import run_in_pipeline

            def _read() -> str:
                reader = PdfReader(str(path))
                if reader.is_encrypted:
                    try:
                        reader.decrypt("")
                    except Exception:
                        return ""
                parts: list[str] = []
                for page in reader.pages:
                    try:
                        parts.append(page.extract_text() or "")
                    except Exception:
                        continue
                return "\n".join(p for p in parts if p)

            text = await asyncio.wait_for(run_in_pipeline(_read), timeout=60)
            text = (text or "").strip()
            return text if len(text) > 10 else None
        except asyncio.TimeoutError:
            logger.warning("pypdf extraction timed out", path=str(path))
            return None
        except Exception as e:
            logger.debug("pypdf extraction failed", path=str(path), error=str(e))
            return None

    async def _try_xlsx_extraction(self, path: Path) -> str | None:
        """Last-resort XLSX text scrape via openpyxl. MarkItDown's xlsx
        converter sometimes returns empty (formula-only sheets with no
        cached values, malformed shared-strings tables, encoding quirks);
        openpyxl can usually still list sheet names and read cell text.
        Returns None on any failure — caller falls through to metadata-only.
        openpyxl comes in via MarkItDown[xlsx], no new dep.

        Output is structured per sheet so a user searching for a sheet
        name like "Midterm" matches even when the cells themselves are
        all numbers.
        """
        try:
            from openpyxl import load_workbook
            from cosma_backend.pipeline_executor import run_in_pipeline

            # Per-cell text cap keeps a single pathological cell from
            # eating the entire token budget; per-sheet cap keeps a
            # 1000-row data sheet from drowning everything else.
            MAX_CELL_CHARS = 500
            MAX_SHEET_CHARS = 50_000

            def _read() -> str:
                # data_only=True returns cached formula results instead of
                # the formula expression itself — usually what the user
                # wants to search by. read_only streams rows, so we don't
                # load a 100MB workbook into memory.
                wb = load_workbook(
                    str(path), data_only=True, read_only=True
                )
                parts: list[str] = []
                for sheet_name in wb.sheetnames:
                    parts.append(f"# Sheet: {sheet_name}")
                    ws = wb[sheet_name]
                    sheet_chars = 0
                    for row in ws.iter_rows(values_only=True):
                        cells = []
                        for value in row:
                            if value is None:
                                continue
                            text = str(value).strip()
                            if not text:
                                continue
                            if len(text) > MAX_CELL_CHARS:
                                text = text[:MAX_CELL_CHARS] + "…"
                            cells.append(text)
                        if cells:
                            line = "\t".join(cells)
                            parts.append(line)
                            sheet_chars += len(line)
                            if sheet_chars > MAX_SHEET_CHARS:
                                parts.append("…(sheet truncated)")
                                break
                wb.close()
                return "\n".join(parts)

            text = await asyncio.wait_for(run_in_pipeline(_read), timeout=60)
            text = (text or "").strip()
            return text if len(text) > 10 else None
        except asyncio.TimeoutError:
            logger.warning("openpyxl extraction timed out", path=str(path))
            return None
        except Exception as e:
            logger.debug("openpyxl extraction failed", path=str(path), error=str(e))
            return None

    async def _try_zip_listing_extraction(self, path: Path) -> str | None:
        """Build a searchable text body for a ZIP from its entry list,
        plus the contents of any small text-like inner files (README,
        .md, .txt, .json, .yaml, manifests, plists). The entry list
        alone is meaningful: `VSCode-darwin-universal.zip` indexed as
        "VSCode.app/Contents/MacOS/Code Helper..." actually surfaces
        the right file when the user searches for "VSCode".
        """
        try:
            from cosma_backend.pipeline_executor import run_in_pipeline

            text_exts = {
                ".txt", ".md", ".markdown", ".rst", ".json", ".yaml",
                ".yml", ".xml", ".plist", ".cfg", ".ini", ".toml",
                ".html", ".htm", ".csv",
            }
            readme_names = {"readme", "license", "notice", "changelog"}
            # Cap how much we scrape so a giant archive doesn't bloat
            # the embedding input. ~120 KB of text ≈ 30k tokens, well
            # under the summarizer's per-chunk budget.
            max_inner_bytes = 120 * 1024
            max_entries_listed = 800

            def _walk() -> str:
                with zipfile.ZipFile(path) as zf:
                    names = zf.namelist()
                    listing = "\n".join(names[:max_entries_listed])
                    if len(names) > max_entries_listed:
                        listing += f"\n... and {len(names) - max_entries_listed} more entries"

                    inner_chunks: list[str] = []
                    remaining = max_inner_bytes
                    for info in zf.infolist():
                        if remaining <= 0 or info.is_dir():
                            continue
                        name = info.filename
                        ext = Path(name).suffix.lower()
                        stem = Path(name).stem.lower()
                        if not (ext in text_exts or stem in readme_names):
                            continue
                        if info.file_size > remaining:
                            continue
                        try:
                            with zf.open(info) as f:
                                raw = f.read(remaining)
                            chunk = raw.decode("utf-8", errors="replace")
                            inner_chunks.append(f"--- {name} ---\n{chunk}")
                            remaining -= len(raw)
                        except Exception:
                            continue
                    body = (
                        f"Archive: {path.name}\n"
                        f"Entries ({len(names)}):\n{listing}\n"
                    )
                    if inner_chunks:
                        body += "\nInner text files:\n" + "\n".join(inner_chunks)
                    return body

            text = await asyncio.wait_for(run_in_pipeline(_walk), timeout=30)
            return text if text and len(text.strip()) > 10 else None
        except asyncio.TimeoutError:
            logger.warning("ZIP listing timed out", path=str(path))
            return None
        except zipfile.BadZipFile:
            logger.debug("ZIP is malformed; listing skipped", path=str(path))
            return None
        except Exception as e:
            logger.debug("ZIP listing failed", path=str(path), error=str(e))
            return None

    def set_extraction_strategy(self, strategy: str) -> None:
        """
        Set extraction strategy.
        
        Args:
            strategy: Strategy to use ('spotlight_first', 'markitdown_only', 'spotlight_only')
        """
        valid_strategies = ["spotlight_first", "markitdown_only", "spotlight_only"]
        if strategy not in valid_strategies:
            raise ValueError(f"Invalid strategy. Must be one of: {valid_strategies}")
            
        self.extraction_strategy = strategy
        logger.info("Extraction strategy updated", strategy=strategy)

    def get_extraction_stats(self) -> dict[str, int]:
        """
        Get statistics on which extraction method was used.
        
        Returns:
            Dictionary with extraction statistics
        """
        return self.extraction_stats.copy()

    def reset_extraction_stats(self) -> None:
        """Reset extraction statistics."""
        for key in self.extraction_stats:
            self.extraction_stats[key] = 0
        logger.info("Extraction statistics reset")


def _human_size(num_bytes: int) -> str:
    """Format a byte count as a human-readable string (e.g. "12.4 MB")."""
    size = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024
    return f"{int(num_bytes)} B"


def _audio_metadata_placeholder(path: Path) -> str:
    """Build a minimal text body for an audio file when transcription fails.

    Indexes the file by name, extension, and size so it surfaces in search
    even when whisper produced no usable transcript (silent track, codec
    not understood, file too short).
    """
    try:
        size = path.stat().st_size
        size_str = _human_size(size)
    except OSError:
        size_str = "unknown size"

    try:
        from cosma_backend.parser import whisper_local
        whisper_present = whisper_local.is_available()
    except Exception:
        whisper_present = True

    reason = (
        "no usable speech detected (silent track, music-only, or speech below the model's confidence threshold)"
        if whisper_present
        else "local whisper backend not installed"
    )

    return (
        f"Audio file: {path.name}\n"
        f"Type: {path.suffix.lower().lstrip('.')} audio, {size_str}.\n"
        f"No speech transcript could be extracted ({reason})."
    )


def _generic_metadata_placeholder(path: Path, reason: str) -> str:
    """Build a minimal text body for a file when every parser came back
    empty. Indexes the file by name, extension, and size so it remains
    findable in semantic search even though we have no real content.
    Mirrors the audio/video placeholders so the metadata-only path is
    consistent across formats.
    """
    try:
        size_str = _human_size(path.stat().st_size)
    except OSError:
        size_str = "unknown size"
    return (
        f"File: {path.name}\n"
        f"Type: {path.suffix.lower().lstrip('.') or 'unknown'}, {size_str}.\n"
        f"No textual content could be extracted ({reason})."
    )


def _video_metadata_placeholder(path: Path) -> str:
    """Build a minimal text body for a video file when both frame and
    audio extraction failed. See ``_audio_metadata_placeholder`` for why
    we degrade to metadata instead of failing the file.

    Checks ffmpeg availability via the resolver so the placeholder can
    distinguish "ffmpeg missing entirely" from "ffmpeg present but the
    codec defeated us" — the second case used to be the only diagnostic
    we surfaced, which led users to think their iPhone HEVC files had
    a codec issue when in reality the backend just couldn't find ffmpeg.
    """
    try:
        size = path.stat().st_size
        size_str = _human_size(size)
    except OSError:
        size_str = "unknown size"

    try:
        from cosma_backend.parser.ffmpeg_runtime import have_system_ffmpeg
        from cosma_backend.parser import whisper_local
        ffmpeg_present = have_system_ffmpeg()
        # The bundled imageio-ffmpeg fallback covers the no-system case,
        # so "ffmpeg present" here means *something* is on disk; we
        # surface the system-vs-bundled distinction only in logs.
        whisper_present = whisper_local.is_available()
    except Exception:
        ffmpeg_present = True
        whisper_present = True

    reasons: list[str] = []
    if not whisper_present:
        reasons.append("local whisper not installed (transcription unavailable)")
    if ffmpeg_present:
        reasons.append("codec unreadable even after transcoding to a clean H.264/AAC proxy")
    else:
        reasons.append("no ffmpeg available — install via Homebrew (`brew install ffmpeg`) for full media support")
    detail = "; ".join(reasons)

    return (
        f"Video file: {path.name}\n"
        f"Type: {path.suffix.lower().lstrip('.')} video, {size_str}.\n"
        f"No transcript or frames could be extracted ({detail})."
    )


def get_supported_extensions() -> set[str]:
    """Get the set of supported file extensions."""
    parser = FileParser()
    return parser.get_supported_extensions()


def is_supported_file(file: File) -> bool:
    """Check if a file format is supported."""
    parser = FileParser()
    return parser.is_supported(file)
